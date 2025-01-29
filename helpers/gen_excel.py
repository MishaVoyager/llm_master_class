import random
import string

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def main():
    with open_excel("something.xlsx") as ws:
        for i in range(30):
            fill_xls(ws, f"Товарчик{i}", 1000, "Произведен в РФ")


def fill_xls(ws: Worksheet, product_name: str, rows: int, value):
    for i in range(rows):
        ws.append([product_name, _gen_random_string(30), value])
    return ws


def add_header(ws: Worksheet):
    ws.append(["Наименование товара", "Код маркировки", "Тип эмиссии"])
    return ws


class open_excel:
    def __init__(self, name: str, append_to_existed_file: bool = False):
        self.name = name
        self.existed_file = append_to_existed_file

    def __enter__(self):
        if self.existed_file:
            wb = load_workbook(self.name)
        else:
            wb = Workbook()
        self.wb = wb
        self.ws = wb.active
        return self.ws

    def __exit__(self, exc_type, exc_value, traceback):
        self.wb.save(self.name)


def _gen_random_string(length: int) -> str:
    return "".join(random.choice(string.ascii_lowercase) for _ in range(length))
